import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Скиллы которые уже есть в нашей библиотеке
already_have = [
    "make-no-mistakes", "test-driven-development", "systematic-debugging",
    "writing-plans", "frontend-design", "brainstorming", "executing-plans",
    "verification-before-completion", "finishing-a-development-branch",
    "skill-creator", "mcp-builder", "webapp-testing", "playwright-automation",
    "semgrep-security-scan", "algorithmic-art", "canvas-design", "gif-creator",
    "ios-simulator"
]

# Все скиллы из репозиториев
skills_data = [
    # ===== JEFFALLAN/CLAUDE-SKILLS (66 скиллов) =====
    # Languages (14)
    {"name": "python-pro", "description": "Python 3.11+ с type hints, async, pytest, mypy", "category": "Language", "source": "jeffallan", "take": "ДА", "reason": "Универсальный, используем Python часто"},
    {"name": "typescript-pro", "description": "TypeScript advanced types, generics, utility types, tRPC", "category": "Language", "source": "jeffallan", "take": "ДА", "reason": "Основной язык для фронтенда"},
    {"name": "javascript-pro", "description": "Modern ES2023+, async patterns, Node.js", "category": "Language", "source": "jeffallan", "take": "НЕТ", "reason": "Покрывается typescript-pro"},
    {"name": "golang-pro", "description": "Go concurrency, goroutines, channels, gRPC", "category": "Language", "source": "jeffallan", "take": "ДА", "reason": "Используем для backend сервисов"},
    {"name": "rust-engineer", "description": "Rust ownership, lifetimes, async с tokio", "category": "Language", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Пока не используем Rust активно"},
    {"name": "swift-expert", "description": "Swift 5.9+, SwiftUI, Combine, async/await", "category": "Language", "source": "jeffallan", "take": "ДА", "reason": "Планируем iOS разработку"},
    {"name": "kotlin-specialist", "description": "Kotlin coroutines, Flow, Jetpack Compose", "category": "Language", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Android - второй приоритет после iOS"},
    {"name": "java-architect", "description": "Enterprise Java, Spring ecosystem, design patterns", "category": "Language", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем Java"},
    {"name": "csharp-developer", "description": "C# .NET 8+, Entity Framework, LINQ, async/await", "category": "Language", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем .NET"},
    {"name": "cpp-pro", "description": "Modern C++20/23, templates, SIMD optimization", "category": "Language", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем C++"},
    {"name": "php-pro", "description": "PHP 8.3+, Laravel patterns, Swoole", "category": "Language", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем PHP"},
    {"name": "sql-pro", "description": "Advanced SQL, window functions, CTEs, optimization", "category": "Language", "source": "jeffallan", "take": "ДА", "reason": "SQL используем везде"},

    # Backend (7)
    {"name": "nestjs-expert", "description": "NestJS модули, DI, TypeORM/Prisma", "category": "Backend", "source": "jeffallan", "take": "ДА", "reason": "Основной backend фреймворк"},
    {"name": "django-expert", "description": "Django, DRF, ORM optimization", "category": "Backend", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Альтернатива для Python backend"},
    {"name": "fastapi-expert", "description": "FastAPI async, Pydantic V2, SQLAlchemy", "category": "Backend", "source": "jeffallan", "take": "ДА", "reason": "Используем для API сервисов"},
    {"name": "spring-boot-engineer", "description": "Spring Boot 3.x, Spring Security, WebFlux", "category": "Backend", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем Java/Spring"},
    {"name": "laravel-specialist", "description": "Laravel 10+, Eloquent, Livewire, Sanctum", "category": "Backend", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем PHP/Laravel"},
    {"name": "rails-expert", "description": "Rails 7+, Hotwire, Turbo, Sidekiq", "category": "Backend", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем Ruby/Rails"},
    {"name": "dotnet-core-expert", "description": ".NET 8 minimal APIs, clean architecture", "category": "Backend", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем .NET"},

    # Frontend (8)
    {"name": "react-expert", "description": "React 18+, hooks, Server Components, Suspense", "category": "Frontend", "source": "jeffallan", "take": "ДА", "reason": "Основной frontend фреймворк"},
    {"name": "nextjs-developer", "description": "Next.js 14+ App Router, Server Actions", "category": "Frontend", "source": "jeffallan", "take": "ДА", "reason": "Используем Next.js"},
    {"name": "vue-expert", "description": "Vue 3 Composition API, Pinia, TypeScript", "category": "Frontend", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Альтернатива React, низкий приоритет"},
    {"name": "vue-expert-js", "description": "Vue 3 с JavaScript only, JSDoc typing", "category": "Frontend", "source": "jeffallan", "take": "НЕТ", "reason": "Предпочитаем TypeScript"},
    {"name": "angular-architect", "description": "Angular 17+ standalone components, signals", "category": "Frontend", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем Angular"},
    {"name": "react-native-expert", "description": "React Native, Expo, navigation, native modules", "category": "Frontend", "source": "jeffallan", "take": "ДА", "reason": "Для кросс-платформенных приложений"},
    {"name": "flutter-expert", "description": "Flutter 3+, Riverpod/Bloc, GoRouter", "category": "Frontend", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Альтернатива React Native"},

    # Infrastructure (6)
    {"name": "kubernetes-specialist", "description": "K8s deployments, Helm, RBAC, NetworkPolicies", "category": "Infrastructure", "source": "jeffallan", "take": "ДА", "reason": "Используем K8s для деплоя"},
    {"name": "terraform-engineer", "description": "Terraform IaC, multi-cloud, modules", "category": "Infrastructure", "source": "jeffallan", "take": "ДА", "reason": "Используем Terraform"},
    {"name": "cloud-architect", "description": "AWS/Azure/GCP architecture, Well-Architected", "category": "Infrastructure", "source": "jeffallan", "take": "ДА", "reason": "Работаем с облаками"},
    {"name": "postgres-pro", "description": "PostgreSQL optimization, replication, JSONB", "category": "Infrastructure", "source": "jeffallan", "take": "ДА", "reason": "Основная БД"},
    {"name": "database-optimizer", "description": "DB performance tuning, indexing, query plans", "category": "Infrastructure", "source": "jeffallan", "take": "ДА", "reason": "Важно для производительности"},

    # API & Architecture (7)
    {"name": "graphql-architect", "description": "GraphQL schema design, Apollo Federation", "category": "API", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Пока используем REST"},
    {"name": "api-designer", "description": "REST API design, OpenAPI, versioning", "category": "API", "source": "jeffallan", "take": "ДА", "reason": "Проектируем REST API"},
    {"name": "websocket-engineer", "description": "WebSockets, Socket.IO, real-time", "category": "API", "source": "jeffallan", "take": "ДА", "reason": "Для real-time функций"},
    {"name": "microservices-architect", "description": "Microservices patterns, DDD, saga, event sourcing", "category": "API", "source": "jeffallan", "take": "ДА", "reason": "Архитектура микросервисов"},
    {"name": "mcp-developer", "description": "MCP protocol, TypeScript/Python SDKs", "category": "API", "source": "jeffallan", "take": "УЖЕ ЕСТЬ", "reason": "Уже есть mcp-builder"},
    {"name": "architecture-designer", "description": "System design, ADRs, scalability", "category": "API", "source": "jeffallan", "take": "ДА", "reason": "Для проектирования систем"},

    # Quality (5)
    {"name": "test-master", "description": "Unit/integration/E2E testing strategy", "category": "Quality", "source": "jeffallan", "take": "ДА", "reason": "Комплексное тестирование"},
    {"name": "playwright-expert", "description": "Playwright E2E, Page Object Model", "category": "Quality", "source": "jeffallan", "take": "УЖЕ ЕСТЬ", "reason": "Уже есть playwright-automation"},
    {"name": "code-reviewer", "description": "PR reviews, code quality audits", "category": "Quality", "source": "jeffallan", "take": "ДА", "reason": "Улучшает качество кода"},
    {"name": "code-documenter", "description": "Docstrings, API docs, OpenAPI specs", "category": "Quality", "source": "jeffallan", "take": "ДА", "reason": "Важно для документации"},
    {"name": "debugging-wizard", "description": "Systematic debugging, root cause analysis", "category": "Quality", "source": "jeffallan", "take": "УЖЕ ЕСТЬ", "reason": "Уже есть systematic-debugging"},

    # DevOps (6)
    {"name": "devops-engineer", "description": "CI/CD pipelines, Docker, GitOps", "category": "DevOps", "source": "jeffallan", "take": "ДА", "reason": "DevOps практики"},
    {"name": "monitoring-expert", "description": "Prometheus/Grafana, logging, tracing", "category": "DevOps", "source": "jeffallan", "take": "ДА", "reason": "Мониторинг важен"},
    {"name": "sre-engineer", "description": "SLIs/SLOs, incident response, reliability", "category": "DevOps", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Когда вырастем"},
    {"name": "chaos-engineer", "description": "Chaos testing, fault injection", "category": "DevOps", "source": "jeffallan", "take": "НЕТ", "reason": "Слишком специфично"},
    {"name": "cli-developer", "description": "CLI tools, argument parsing, prompts", "category": "DevOps", "source": "jeffallan", "take": "ДА", "reason": "Создаём CLI утилиты"},

    # Security (4)
    {"name": "secure-code-guardian", "description": "Secure coding, OWASP Top 10 prevention", "category": "Security", "source": "jeffallan", "take": "ДА", "reason": "Безопасность критична"},
    {"name": "security-reviewer", "description": "Security audits, SAST, penetration testing", "category": "Security", "source": "jeffallan", "take": "ДА", "reason": "Аудит безопасности"},
    {"name": "fullstack-guardian", "description": "Full-stack feature implementation", "category": "Security", "source": "jeffallan", "take": "ДА", "reason": "End-to-end разработка"},

    # Data & ML (7)
    {"name": "pandas-pro", "description": "DataFrame manipulation, data cleaning", "category": "Data/ML", "source": "jeffallan", "take": "ДА", "reason": "Анализ данных"},
    {"name": "spark-engineer", "description": "Apache Spark, PySpark, distributed processing", "category": "Data/ML", "source": "jeffallan", "take": "НЕТ", "reason": "Пока не используем Big Data"},
    {"name": "ml-pipeline", "description": "ML pipelines, MLflow, Kubeflow", "category": "Data/ML", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Когда начнём ML"},
    {"name": "prompt-engineer", "description": "LLM prompts, chain-of-thought, evaluation", "category": "Data/ML", "source": "jeffallan", "take": "ДА", "reason": "Работаем с LLM"},
    {"name": "rag-architect", "description": "RAG systems, vector DBs, embeddings", "category": "Data/ML", "source": "jeffallan", "take": "ДА", "reason": "Строим RAG системы"},
    {"name": "fine-tuning-expert", "description": "LLM fine-tuning, LoRA, PEFT", "category": "Data/ML", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Для продвинутого ML"},

    # Platform (4)
    {"name": "salesforce-developer", "description": "Apex, LWC, SOQL, governor limits", "category": "Platform", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем Salesforce"},
    {"name": "shopify-expert", "description": "Liquid, Storefront API, checkout", "category": "Platform", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем Shopify"},
    {"name": "wordpress-pro", "description": "WordPress themes, plugins, Gutenberg", "category": "Platform", "source": "jeffallan", "take": "НЕТ", "reason": "Не используем WordPress"},
    {"name": "atlassian-mcp", "description": "Jira/Confluence integration via MCP", "category": "Platform", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Может пригодиться"},

    # Specialized (3)
    {"name": "legacy-modernizer", "description": "Legacy code modernization, strangler fig", "category": "Specialized", "source": "jeffallan", "take": "ПОЗЖЕ", "reason": "Для legacy проектов"},
    {"name": "embedded-systems", "description": "Firmware, RTOS, STM32, ESP32", "category": "Specialized", "source": "jeffallan", "take": "НЕТ", "reason": "Не делаем embedded"},
    {"name": "game-developer", "description": "Unity, Unreal, game patterns", "category": "Specialized", "source": "jeffallan", "take": "НЕТ", "reason": "Не делаем игры"},

    # Workflow (3)
    {"name": "feature-forge", "description": "Feature definition, requirements, specs", "category": "Workflow", "source": "jeffallan", "take": "ДА", "reason": "Для определения фич"},
    {"name": "spec-miner", "description": "Reverse-engineering specs from code", "category": "Workflow", "source": "jeffallan", "take": "ДА", "reason": "Для анализа legacy"},
    {"name": "the-fool", "description": "Devil's advocate, pre-mortem, red team", "category": "Workflow", "source": "jeffallan", "take": "ДА", "reason": "Критическое мышление"},

    # ===== TRAIL OF BITS (28 скиллов) =====
    {"name": "ask-questions-if-underspecified", "description": "Задаёт уточняющие вопросы перед работой", "category": "Workflow", "source": "trailofbits", "take": "ДА", "reason": "Улучшает качество работы"},
    {"name": "audit-context-building", "description": "Глубокий архитектурный контекст для аудита", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Для security аудита"},
    {"name": "building-secure-contracts", "description": "Smart contract security toolkit, 6 сканеров", "category": "Security", "source": "trailofbits", "take": "НЕТ", "reason": "Не делаем blockchain"},
    {"name": "burpsuite-project-parser", "description": "Парсинг Burp Suite проектов", "category": "Security", "source": "trailofbits", "take": "НЕТ", "reason": "Слишком специфично"},
    {"name": "claude-in-chrome-troubleshooting", "description": "Диагностика Chrome MCP extension", "category": "Workflow", "source": "trailofbits", "take": "НЕТ", "reason": "Слишком специфично"},
    {"name": "constant-time-analysis", "description": "Timing side-channel vulnerabilities", "category": "Security", "source": "trailofbits", "take": "НЕТ", "reason": "Криптография не наш фокус"},
    {"name": "culture-index", "description": "Интерпретация Culture Index опросов", "category": "Workflow", "source": "trailofbits", "take": "НЕТ", "reason": "HR инструмент"},
    {"name": "debug-buttercup", "description": "Debug Buttercup CRS на K8s", "category": "Workflow", "source": "trailofbits", "take": "НЕТ", "reason": "Специфичный продукт"},
    {"name": "devcontainer-setup", "description": "Pre-configured devcontainers с Claude Code", "category": "Workflow", "source": "trailofbits", "take": "ДА", "reason": "Полезно для onboarding"},
    {"name": "differential-review", "description": "Security-focused diff review", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Security review изменений"},
    {"name": "dwarf-expert", "description": "Анализ DWARF debug files", "category": "Security", "source": "trailofbits", "take": "НЕТ", "reason": "Слишком низкоуровнево"},
    {"name": "entry-point-analyzer", "description": "Entry points в smart contracts", "category": "Security", "source": "trailofbits", "take": "НЕТ", "reason": "Не делаем blockchain"},
    {"name": "firebase-apk-scanner", "description": "Firebase security в APK", "category": "Security", "source": "trailofbits", "take": "ПОЗЖЕ", "reason": "Когда делаем Android"},
    {"name": "gh-cli", "description": "GitHub CLI интеграция", "category": "Workflow", "source": "trailofbits", "take": "ДА", "reason": "Полезно для GitHub"},
    {"name": "git-cleanup", "description": "Cleanup worktrees и branches", "category": "Workflow", "source": "trailofbits", "take": "ДА", "reason": "Полезно для git hygiene"},
    {"name": "insecure-defaults", "description": "Hardcoded secrets, weak crypto", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Поиск уязвимостей"},
    {"name": "modern-python", "description": "uv, ruff, pytest best practices", "category": "Workflow", "source": "trailofbits", "take": "ДА", "reason": "Modern Python tooling"},
    {"name": "property-based-testing", "description": "Property-based testing guidance", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Продвинутое тестирование"},
    {"name": "second-opinion", "description": "Code review через другие LLM", "category": "Workflow", "source": "trailofbits", "take": "ДА", "reason": "Дополнительная проверка"},
    {"name": "semgrep-rule-creator", "description": "Создание Semgrep rules", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Custom security rules"},
    {"name": "semgrep-rule-variant-creator", "description": "Порт Semgrep rules на другие языки", "category": "Security", "source": "trailofbits", "take": "ПОЗЖЕ", "reason": "После semgrep-rule-creator"},
    {"name": "sharp-edges", "description": "Error-prone APIs, dangerous configs", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Поиск footguns"},
    {"name": "spec-to-code-compliance", "description": "Spec-to-code verification", "category": "Security", "source": "trailofbits", "take": "НЕТ", "reason": "Blockchain focused"},
    {"name": "static-analysis", "description": "CodeQL, Semgrep, SARIF toolkit", "category": "Security", "source": "trailofbits", "take": "УЖЕ ЕСТЬ", "reason": "Уже есть semgrep-security-scan"},
    {"name": "testing-handbook-skills", "description": "Trail of Bits testing handbook", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Security testing practices"},
    {"name": "variant-analysis", "description": "Поиск похожих уязвимостей", "category": "Security", "source": "trailofbits", "take": "ДА", "reason": "Поиск паттернов багов"},
    {"name": "workflow-skill-design", "description": "Design patterns для skills", "category": "Workflow", "source": "trailofbits", "take": "ДА", "reason": "Для создания своих skills"},
    {"name": "yara-authoring", "description": "YARA-X detection rules", "category": "Security", "source": "trailofbits", "take": "НЕТ", "reason": "Malware detection не наш фокус"},

    # ===== ANTHROPICS/SKILLS (16 скиллов) =====
    {"name": "algorithmic-art", "description": "Generative art на p5.js", "category": "Creative", "source": "anthropics", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "brand-guidelines", "description": "Anthropic brand colors и typography", "category": "Creative", "source": "anthropics", "take": "НЕТ", "reason": "Специфично для Anthropic"},
    {"name": "canvas-design", "description": "Visual art в PNG/PDF", "category": "Creative", "source": "anthropics", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "doc-coauthoring", "description": "Co-authoring документов", "category": "Docs", "source": "anthropics", "take": "ДА", "reason": "Совместная работа над доками"},
    {"name": "docx", "description": "Создание/редактирование Word", "category": "Docs", "source": "anthropics", "take": "ДА", "reason": "Работа с документами"},
    {"name": "frontend-design", "description": "Distinctive UI design", "category": "Creative", "source": "anthropics", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "internal-comms", "description": "Internal communications", "category": "Docs", "source": "anthropics", "take": "НЕТ", "reason": "Корпоративное"},
    {"name": "mcp-builder", "description": "MCP servers guide", "category": "API", "source": "anthropics", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "pdf", "description": "PDF manipulation toolkit", "category": "Docs", "source": "anthropics", "take": "ДА", "reason": "Работа с PDF"},
    {"name": "pptx", "description": "PowerPoint presentations", "category": "Docs", "source": "anthropics", "take": "ДА", "reason": "Презентации"},
    {"name": "skill-creator", "description": "Guide для создания skills", "category": "Meta", "source": "anthropics", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "slack-gif-creator", "description": "Animated GIFs для Slack", "category": "Creative", "source": "anthropics", "take": "УЖЕ ЕСТЬ", "reason": "Уже есть gif-creator"},
    {"name": "theme-factory", "description": "Professional themes toolkit", "category": "Creative", "source": "anthropics", "take": "ДА", "reason": "Темы для артефактов"},
    {"name": "web-artifacts-builder", "description": "Complex HTML artifacts", "category": "Creative", "source": "anthropics", "take": "ДА", "reason": "Для сложных артефактов"},
    {"name": "webapp-testing", "description": "Playwright testing toolkit", "category": "Testing", "source": "anthropics", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "xlsx", "description": "Excel spreadsheets", "category": "Docs", "source": "anthropics", "take": "ДА", "reason": "Работа с таблицами"},

    # ===== OBRA/SUPERPOWERS (14 скиллов) =====
    {"name": "brainstorming", "description": "Discovery диалог перед дизайном", "category": "Workflow", "source": "obra", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "dispatching-parallel-agents", "description": "Параллельные subagents", "category": "Workflow", "source": "obra", "take": "ДА", "reason": "Параллельная работа"},
    {"name": "executing-plans", "description": "Batch execution с checkpoints", "category": "Workflow", "source": "obra", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "finishing-a-development-branch", "description": "Merge/PR workflow", "category": "Workflow", "source": "obra", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "receiving-code-review", "description": "Обработка code review feedback", "category": "Workflow", "source": "obra", "take": "ДА", "reason": "Работа с ревью"},
    {"name": "requesting-code-review", "description": "Запрос code review", "category": "Workflow", "source": "obra", "take": "ДА", "reason": "Автоматизация ревью"},
    {"name": "subagent-driven-development", "description": "Two-stage review per task", "category": "Workflow", "source": "obra", "take": "ДА", "reason": "Quality gates"},
    {"name": "systematic-debugging", "description": "4-фазный debugging", "category": "Workflow", "source": "obra", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "test-driven-development", "description": "TDD методология", "category": "Workflow", "source": "obra", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "using-git-worktrees", "description": "Isolated git worktrees", "category": "Workflow", "source": "obra", "take": "ДА", "reason": "Параллельная работа"},
    {"name": "using-superpowers", "description": "How to use skills", "category": "Meta", "source": "obra", "take": "НЕТ", "reason": "Специфично для superpowers"},
    {"name": "verification-before-completion", "description": "Verification перед completion", "category": "Workflow", "source": "obra", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "writing-plans", "description": "Детальные планы", "category": "Workflow", "source": "obra", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "writing-skills", "description": "Создание новых skills", "category": "Meta", "source": "obra", "take": "ДА", "reason": "Для создания своих skills"},

    # ===== COMMUNITY SKILLS =====
    {"name": "ios-simulator-skill", "description": "21 скрипт для iOS автоматизации", "category": "Mobile", "source": "conorluddy", "take": "УЖЕ ЕСТЬ", "reason": "Уже добавили"},
    {"name": "dev-browser", "description": "Browser для Claude agents", "category": "Tools", "source": "sawyerhood", "take": "ДА", "reason": "Browser automation"},
    {"name": "claude-d3js-skill", "description": "D3.js визуализации", "category": "Creative", "source": "chrisvoncsefalvay", "take": "ДА", "reason": "Data visualization"},
]

# Создаём DataFrame
df = pd.DataFrame(skills_data)

# Сортируем по категории и решению
df = df.sort_values(['take', 'category', 'name'])

# Создаём Excel файл
wb = Workbook()
ws = wb.active
ws.title = "Skills Analysis"

# Заголовки
headers = ["Название", "Описание", "Категория", "Источник", "Берём?", "Почему"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Данные
colors = {
    "ДА": "C6EFCE",      # Зелёный
    "НЕТ": "FFC7CE",     # Красный
    "ПОЗЖЕ": "FFEB9C",   # Жёлтый
    "УЖЕ ЕСТЬ": "B4C6E7" # Синий
}

for row_idx, row_data in enumerate(df.values, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Цвет для колонки "Берём?"
        if col_idx == 5:  # take column
            if value in colors:
                cell.fill = PatternFill(start_color=colors[value], end_color=colors[value], fill_type="solid")

# Ширина колонок
ws.column_dimensions['A'].width = 30  # Название
ws.column_dimensions['B'].width = 50  # Описание
ws.column_dimensions['C'].width = 15  # Категория
ws.column_dimensions['D'].width = 15  # Источник
ws.column_dimensions['E'].width = 12  # Берём?
ws.column_dimensions['F'].width = 40  # Почему

# Сохраняем
output_path = "/Users/alik/project/claudecode/skills/skills_analysis.xlsx"
wb.save(output_path)
print(f"Excel создан: {output_path}")

# Статистика
print(f"\nСтатистика:")
print(f"Всего скиллов: {len(df)}")
print(f"ДА: {len(df[df['take'] == 'ДА'])}")
print(f"НЕТ: {len(df[df['take'] == 'НЕТ'])}")
print(f"ПОЗЖЕ: {len(df[df['take'] == 'ПОЗЖЕ'])}")
print(f"УЖЕ ЕСТЬ: {len(df[df['take'] == 'УЖЕ ЕСТЬ'])}")
